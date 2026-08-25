#!/usr/bin/env python3
"""
Convierte el "Reporte de cortes" en Excel de la Junta Rural de Agua y Saneamiento de
Col. Hidalgo en SQL listo para Supabase.

Uso:
    python3 scripts/generar_seed.py "/ruta/REZAGO AGOSTO 2026.xlsx"

Genera dentro de supabase/datos_privados/ (carpeta ignorada por git porque el
padron contiene datos personales de los cuentahabientes):

    datos_padron.sql    solo los UPSERT del padron
    setup_completo.sql  esquema desde cero + padron, para pegar en el editor
                        SQL de un proyecto nuevo de Supabase

Columnas esperadas en la hoja (encabezado en la fila 5):
    IdUsuario | Cuenta | Nombre | Direccion | No. Medidor | Ruta | Sec. |
    Ultimo pago | tarifa | adeudo | Consumo | Observacion
"""

from __future__ import annotations

import datetime as dt
import re
import sys
from pathlib import Path

import openpyxl

RAIZ = Path(__file__).resolve().parent.parent
DESTINO = RAIZ / "supabase" / "datos_privados"
MIGRACIONES = RAIZ / "supabase" / "migrations"
FILA_ENCABEZADO = 5


def sql_texto(valor) -> str:
    if valor is None:
        return "null"
    texto = re.sub(r"\s+", " ", str(valor)).strip()
    if not texto:
        return "null"
    return "'" + texto.replace("'", "''") + "'"


def sql_num(valor) -> str:
    return "null" if valor is None else str(valor)


def sql_fecha(valor) -> str:
    if isinstance(valor, dt.datetime):
        valor = valor.date()
    if not isinstance(valor, dt.date):
        return "null"
    return f"'{valor.isoformat()}'"


def meses_entre(desde, hasta: dt.date) -> int:
    """Meses completos entre el ultimo pago y la fecha de corte."""
    if isinstance(desde, dt.datetime):
        desde = desde.date()
    if not isinstance(desde, dt.date):
        return 0
    meses = (hasta.year - desde.year) * 12 + (hasta.month - desde.month)
    if hasta.day < desde.day:
        meses -= 1
    return max(meses, 0)


def leer(ruta: Path):
    hoja = openpyxl.load_workbook(ruta, data_only=True).worksheets[0]

    fecha_corte = None
    for fila in hoja.iter_rows(min_row=1, max_row=FILA_ENCABEZADO, values_only=True):
        if fila and str(fila[0] or "").strip().lower() == "fecha":
            if isinstance(fila[1], dt.datetime):
                fecha_corte = fila[1].date()
            elif isinstance(fila[1], dt.date):
                fecha_corte = fila[1]
    fecha_corte = fecha_corte or dt.date.today()

    registros = []
    for fila in hoja.iter_rows(min_row=FILA_ENCABEZADO + 1, values_only=True):
        id_usuario, cuenta, nombre = fila[0], fila[1], fila[2]
        # La ultima fila del reporte es el marcador "FIN DEL REPORTE".
        if not isinstance(id_usuario, int) or not cuenta or not nombre:
            continue
        registros.append(
            {
                "id": f"cta-{id_usuario}",
                "id_usuario": id_usuario,
                "numero_cuenta": str(cuenta).strip(),
                "nombre": nombre,
                "direccion": fila[3],
                "no_medidor": fila[4],
                "ruta": fila[5],
                "secuencia": fila[6],
                "ultimo_pago": fila[7],
                "tarifa": fila[8],
                "saldo_vencido": round(float(fila[9] or 0), 2),
                "consumo": fila[10],
                "meses_adeudo": meses_entre(fila[7], fecha_corte),
                "fecha_corte": fecha_corte,
            }
        )
    return registros, fecha_corte


COLUMNAS = (
    "id, id_usuario, numero_cuenta, nombre, direccion, no_medidor, ruta, "
    "secuencia, ultimo_pago, tarifa, saldo_vencido, meses_adeudo, consumo, "
    "fecha_corte"
)


def bloque_datos(registros, fecha_corte) -> str:
    valores = []
    for r in registros:
        valores.append(
            "  ("
            + ", ".join(
                [
                    sql_texto(r["id"]),
                    sql_num(r["id_usuario"]),
                    sql_texto(r["numero_cuenta"]),
                    sql_texto(r["nombre"]),
                    sql_texto(r["direccion"]),
                    sql_texto(r["no_medidor"]),
                    sql_num(r["ruta"]),
                    sql_num(r["secuencia"]),
                    sql_fecha(r["ultimo_pago"]),
                    sql_texto(r["tarifa"]),
                    f"{r['saldo_vencido']:.2f}",
                    str(r["meses_adeudo"]),
                    sql_num(r["consumo"]),
                    sql_fecha(r["fecha_corte"]),
                ]
            )
            + ")"
        )

    total = sum(r["saldo_vencido"] for r in registros)
    filas = ",\n".join(valores)
    return f"""-- ============================================================================
-- Padron de rezago - Junta Rural de Agua y Saneamiento de Col. Hidalgo
-- Fecha de corte: {fecha_corte.isoformat()}   Cuentas: {len(registros)}   Adeudo: $ {total:,.2f} MXN
--
-- Generado por scripts/generar_seed.py a partir del reporte de cortes en Excel.
-- No editar a mano: vuelve a correr el script con el Excel actualizado.
--
-- Es idempotente: se puede volver a ejecutar cada mes. Empata por numero de
-- cuenta, actualiza saldo/consumo/ultimo pago y NO borra convenios existentes.
-- ============================================================================

insert into public.cuentahabientes (
  {COLUMNAS}
) values
{filas}
on conflict (numero_cuenta) do update set
  id_usuario    = excluded.id_usuario,
  nombre        = excluded.nombre,
  direccion     = excluded.direccion,
  no_medidor    = excluded.no_medidor,
  ruta          = excluded.ruta,
  secuencia     = excluded.secuencia,
  ultimo_pago   = excluded.ultimo_pago,
  tarifa        = excluded.tarifa,
  saldo_vencido = excluded.saldo_vencido,
  meses_adeudo  = excluded.meses_adeudo,
  consumo       = excluded.consumo,
  fecha_corte   = excluded.fecha_corte,
  updated_at    = now();
"""


def main() -> int:
    if len(sys.argv) < 2:
        print(__doc__)
        return 1
    ruta = Path(sys.argv[1]).expanduser()
    if not ruta.exists():
        print(f"No se encontro el archivo: {ruta}")
        return 1

    registros, fecha_corte = leer(ruta)
    if not registros:
        print("El reporte no tiene filas de datos.")
        return 1

    DESTINO.mkdir(parents=True, exist_ok=True)
    datos = bloque_datos(registros, fecha_corte)
    (DESTINO / "datos_padron.sql").write_text(datos, encoding="utf-8")

    # setup_completo.sql = todas las migraciones en orden + el padron.
    partes = [
        "-- ============================================================================",
        "-- Instalacion completa en un proyecto de Supabase vacio.",
        "-- Pegar TODO este archivo en el editor SQL del proyecto y ejecutar.",
        "-- ============================================================================",
        "",
    ]
    for archivo in sorted(MIGRACIONES.glob("*.sql")):
        partes.append(f"\n-- ==== {archivo.name} ====")
        partes.append(archivo.read_text(encoding="utf-8"))
    partes.append("\n-- ==== padron de rezago ====")
    partes.append(datos)
    (DESTINO / "setup_completo.sql").write_text("\n".join(partes), encoding="utf-8")

    total = sum(r["saldo_vencido"] for r in registros)
    print(f"Cuentas:      {len(registros)}")
    print(f"Fecha corte:  {fecha_corte}")
    print(f"Adeudo total: $ {total:,.2f} MXN")
    print(f"Escrito en:   {DESTINO}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
